from io import BytesIO
import re
from datetime import UTC, datetime

from openpyxl import Workbook
import pandas as pd


NEW_SKU_FLAG = "КП"
DERIVED_EMPTY_COLUMNS = {
    "Код аналога",
    "Ценовой сегмент",
    "Решение",
    "Т на упаковку",
    "ФМ на упаковку",
    "Продажи в упаковках",
    "Кол-во аптек с буфером",
    "ФМ процент",
    "Т процент",
    "Сумма продажи",
    "Сумма без скидки",
    "Сумма Т",
    "Цена нулевая",
}
PK_COLUMN = "ПК"
SERVICE_SHEET_NAMES = [
    "Сводка",
    "КП_распознано",
    "Сопоставление",
    "Ошибки",
    "Спорные_позиции",
]
RECOGNIZED_COLUMNS = [
    "sku_index",
    "sku_name_raw",
    "sku_name_normalized",
    "mnn",
    "mnn_status",
    "form",
    "dosage",
    "pack_size",
    "producer",
    "concern",
    "zc",
    "rc",
    "sip_price",
    "bonus_rub",
    "bonus_percent",
    "vat_status",
    "source_sheet",
    "source_fragment",
    "agent_comment",
]
MATCH_COLUMNS = [
    "sku_index",
    "sku_name_raw",
    "selected_tg",
    "selected_tk",
    "selected_pk",
    "confidence",
    "status",
    "match_reason",
    "user_selected_pk",
    "clarification_question",
    "final_action",
]
ERROR_COLUMNS = ["error_code", "severity", "message", "recommendation", "source"]
DISPUTED_COLUMNS = [
    "sku_index",
    "sku_name_raw",
    "issue_type",
    "issue_description",
    "candidate_pk_1",
    "candidate_pk_2",
    "candidate_pk_3",
    "user_decision",
    "agent_comment",
]


def build_pk_workbook(
    base_df: pd.DataFrame,
    matched_skus: list[dict[str, object]],
) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_pk_tabs(workbook, base_df, matched_skus)

    if not workbook.sheetnames:
        workbook.create_sheet("Без_ПК")

    return workbook


def build_result_workbook(
    base_df: pd.DataFrame,
    matched_skus: list[dict[str, object]],
    session_id: str,
    warnings: list[str] | None = None,
    errors: list[dict[str, object]] | None = None,
    metadata: dict[str, object] | None = None,
) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    warnings = warnings or []
    errors = errors or []
    metadata = metadata or {}

    _add_summary_sheet(workbook, session_id, matched_skus, warnings, metadata)
    _add_recognized_sheet(workbook, matched_skus)
    _add_matching_sheet(workbook, matched_skus)
    _add_errors_sheet(workbook, errors)
    _add_disputed_sheet(workbook, matched_skus)
    _add_pk_tabs(workbook, base_df, matched_skus)

    return workbook


def _add_pk_tabs(
    workbook: Workbook,
    base_df: pd.DataFrame,
    matched_skus: list[dict[str, object]],
) -> None:
    for pk, items in _group_items_by_pk(matched_skus).items():
        sheet = workbook.create_sheet(_unique_sheet_name(workbook.sheetnames, pk))
        pk_rows = base_df[base_df[PK_COLUMN] == pk] if PK_COLUMN in base_df.columns else base_df.iloc[0:0]
        _append_dataframe(sheet, pk_rows, list(base_df.columns))

        context = _pk_context(pk_rows, pk)
        for item in items:
            sheet.append(_new_sku_row(item, context, list(base_df.columns)))


def build_pk_workbook_bytes(
    base_df: pd.DataFrame,
    matched_skus: list[dict[str, object]],
) -> bytes:
    workbook = build_pk_workbook(base_df, matched_skus)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_result_workbook_bytes(
    base_df: pd.DataFrame,
    matched_skus: list[dict[str, object]],
    session_id: str,
    warnings: list[str] | None = None,
    errors: list[dict[str, object]] | None = None,
    metadata: dict[str, object] | None = None,
) -> bytes:
    workbook = build_result_workbook(base_df, matched_skus, session_id, warnings, errors, metadata)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _add_summary_sheet(
    workbook: Workbook,
    session_id: str,
    matched_skus: list[dict[str, object]],
    warnings: list[str],
    metadata: dict[str, object],
) -> None:
    sheet = workbook.create_sheet("Сводка")
    sheet.append(["field", "value"])
    pk_count = len(_group_items_by_pk(matched_skus))
    status_counts = _status_counts(matched_skus)
    rows = {
        "created_at": datetime.now(UTC).isoformat(),
        "session_id": session_id,
        "base_filename": metadata.get("base_filename"),
        "offer_filename": metadata.get("offer_filename"),
        "plain_text_used": metadata.get("plain_text_used", False),
        "skus_total": len(matched_skus),
        "pk_tabs_created": pk_count,
        "auto_matched": status_counts.get("auto_matched", 0),
        "review_required": status_counts.get("review_required", 0),
        "need_clarification": status_counts.get("need_clarification", 0),
        "unmatched": status_counts.get("unmatched", 0),
        "warnings": "; ".join(warnings),
        "status": "ready",
    }
    for key, value in rows.items():
        sheet.append([key, value])


def _add_recognized_sheet(
    workbook: Workbook,
    matched_skus: list[dict[str, object]],
) -> None:
    sheet = workbook.create_sheet("КП_распознано")
    sheet.append(RECOGNIZED_COLUMNS)
    for index, item in enumerate(matched_skus):
        sku = _item_sku(item)
        sheet.append(
            [
                index,
                sku.get("sku_name_raw"),
                sku.get("sku_name_normalized"),
                sku.get("mnn"),
                sku.get("mnn_status"),
                sku.get("form"),
                sku.get("dosage"),
                sku.get("pack_size"),
                sku.get("producer"),
                sku.get("concern"),
                sku.get("zc"),
                sku.get("rc"),
                sku.get("sip"),
                sku.get("bonus_per_pack"),
                sku.get("bonus_percent"),
                _vat_status(sku),
                sku.get("source_sheet"),
                sku.get("source_fragment"),
                sku.get("agent_comment"),
            ]
        )


def _add_matching_sheet(
    workbook: Workbook,
    matched_skus: list[dict[str, object]],
) -> None:
    sheet = workbook.create_sheet("Сопоставление")
    sheet.append(MATCH_COLUMNS)
    for index, item in enumerate(matched_skus):
        sku = _item_sku(item)
        match = _item_match(item)
        sheet.append(
            [
                index,
                sku.get("sku_name_raw"),
                match.get("tg"),
                match.get("tk"),
                match.get("user_selected_pk") or match.get("pk"),
                match.get("confidence"),
                match.get("status"),
                match.get("reason"),
                match.get("user_selected_pk"),
                match.get("clarification_question"),
                match.get("reason") if match.get("status") in {"excluded", "unmatched"} else None,
            ]
        )


def _add_errors_sheet(
    workbook: Workbook,
    errors: list[dict[str, object]],
) -> None:
    sheet = workbook.create_sheet("Ошибки")
    sheet.append(ERROR_COLUMNS)
    for error in errors:
        sheet.append([error.get(column) for column in ERROR_COLUMNS])


def _add_disputed_sheet(
    workbook: Workbook,
    matched_skus: list[dict[str, object]],
) -> None:
    sheet = workbook.create_sheet("Спорные_позиции")
    sheet.append(DISPUTED_COLUMNS)
    for index, item in enumerate(matched_skus):
        match = _item_match(item)
        if match.get("status") not in {"review_required", "need_clarification", "unmatched"}:
            continue

        sku = _item_sku(item)
        candidates = match.get("candidates")
        candidates = candidates if isinstance(candidates, list) else []
        sheet.append(
            [
                index,
                sku.get("sku_name_raw"),
                match.get("status"),
                match.get("reason"),
                _candidate_pk(candidates, 0),
                _candidate_pk(candidates, 1),
                _candidate_pk(candidates, 2),
                match.get("user_decision"),
                sku.get("agent_comment"),
            ]
        )


def _append_dataframe(
    sheet: object,
    df: pd.DataFrame,
    columns: list[str],
) -> None:
    sheet.append(columns)
    for _, row in df.iterrows():
        sheet.append([_clean_cell(row.get(column)) for column in columns])


def _group_items_by_pk(
    matched_skus: list[dict[str, object]],
) -> dict[str, list[dict[str, object]]]:
    groups: dict[str, list[dict[str, object]]] = {}
    for item in matched_skus:
        if _is_excluded(item):
            continue

        pk = _item_pk(item)
        if pk is None:
            continue

        groups.setdefault(pk, []).append(item)

    return groups


def _new_sku_row(
    item: dict[str, object],
    context: dict[str, object],
    columns: list[str],
) -> list[object]:
    sku = item.get("sku")
    sku = sku if isinstance(sku, dict) else item

    values = {column: None for column in columns}
    values.update(
        {
            "ТК": context.get("ТК"),
            "ТГ": context.get("ТГ"),
            "ПК": context.get("ПК"),
            "МНН": sku.get("mnn"),
            "Концерн": sku.get("concern"),
            "Номенклатура": sku.get("sku_name_raw"),
            "Признак": NEW_SKU_FLAG,
            "РЦ": sku.get("rc"),
            "Бонус процент": sku.get("bonus_percent"),
            "Бонус на упаковку": sku.get("bonus_per_pack"),
            "Цена СИП": sku.get("sip"),
            "ЗЦ": sku.get("zc"),
        }
    )

    for column in DERIVED_EMPTY_COLUMNS:
        if column in values and column not in {"РЦ", "Бонус процент", "Бонус на упаковку", "Цена СИП", "ЗЦ"}:
            values[column] = None

    return [values.get(column) for column in columns]


def _pk_context(pk_rows: pd.DataFrame, pk: str) -> dict[str, object]:
    if pk_rows.empty:
        return {"ПК": pk, "ТК": None, "ТГ": None}

    first_row = pk_rows.iloc[0]
    return {
        "ПК": pk,
        "ТК": _clean_cell(first_row.get("ТК")),
        "ТГ": _clean_cell(first_row.get("ТГ")),
    }


def _item_pk(item: dict[str, object]) -> str | None:
    match = item.get("match")
    if isinstance(match, dict):
        pk = match.get("user_selected_pk") or match.get("pk")
    else:
        pk = item.get("pk")

    if not pk:
        return None

    text = str(pk).strip()
    return text or None


def _item_sku(item: dict[str, object]) -> dict[str, object]:
    sku = item.get("sku")
    return sku if isinstance(sku, dict) else item


def _item_match(item: dict[str, object]) -> dict[str, object]:
    match = item.get("match")
    return match if isinstance(match, dict) else {}


def _candidate_pk(candidates: list[object], index: int) -> object:
    if index >= len(candidates) or not isinstance(candidates[index], dict):
        return None
    return candidates[index].get("pk") or candidates[index].get("pk_name")


def _status_counts(matched_skus: list[dict[str, object]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for item in matched_skus:
        status = _item_match(item).get("status")
        if isinstance(status, str):
            counts[status] = counts.get(status, 0) + 1
    return counts


def _is_excluded(item: dict[str, object]) -> bool:
    match = item.get("match")
    if not isinstance(match, dict):
        return False
    return match.get("status") in {"excluded", "unmatched"} or match.get("excluded") is True


def _clean_cell(value: object) -> object:
    if pd.isna(value):
        return None
    return value


def _vat_status(sku: dict[str, object]) -> object:
    if sku.get("vat_status") is not None:
        return sku.get("vat_status")
    if sku.get("vat_included") is True:
        return "with_vat"
    if sku.get("vat_included") is False:
        return "without_vat"
    return None


def _unique_sheet_name(existing_names: list[str], pk: str) -> str:
    base_name = re.sub(r"[\[\]:*?/\\]", "_", pk).strip() or "ПК"
    base_name = base_name[:31]
    if base_name not in existing_names:
        return base_name

    counter = 2
    while True:
        suffix = f"_{counter}"
        candidate = f"{base_name[: 31 - len(suffix)]}{suffix}"
        if candidate not in existing_names:
            return candidate
        counter += 1
