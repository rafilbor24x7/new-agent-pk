from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app


class RuleBasedLLMClient:
    def can_match_pk(self):
        return True

    def match_pk(self, sku, pk_list, candidates):
        text = " ".join(str(sku.get(field) or "") for field in ("trade_name", "mnn", "atx_code", "ftg_name"))
        if "Ибупрофен" in text or "M01AE" in text:
            return _llm_result(pk_list[0], "ibuprofen_atx")
        if "Парацетамол" in text or "N02BE" in text:
            return _llm_result(pk_list[1], "paracetamol_atx")
        if "Лоратадин" in text or "R06AX" in text:
            return _llm_result(pk_list[2], "loratadine_atx")
        if "Амоксициллин" in text or "J01CA" in text:
            return _llm_result(pk_list[3], "amoxicillin_atx")
        return _llm_result(pk_list[4], "fallback")


def _llm_result(item, reason):
    return {
        "pk": item["pk"],
        "tg": item["tg"],
        "tk": item["tk"],
        "confidence": 0.94,
        "reason": reason,
    }


def test_full_pipeline(monkeypatch, tmp_path):
    from app.api import tools

    monkeypatch.setenv("ESKLP_DIR", "data/esklp_test")
    tools.get_esklp_lookup.cache_clear()
    monkeypatch.setattr(
        tools,
        "load_pk_list",
        lambda: [
            {"tg": "Лекарства", "tk": "НПВП", "pk": "Ибупрофен таблетки"},
            {"tg": "Лекарства", "tk": "Анальгетики", "pk": "Парацетамол таблетки"},
            {"tg": "Лекарства", "tk": "Антигистаминные", "pk": "Лоратадин таблетки"},
            {"tg": "Лекарства", "tk": "Антибиотики", "pk": "Амоксициллин капсулы"},
            {"tg": "Лекарства", "tk": "Анальгетики", "pk": "Цитрамон таблетки"},
        ],
    )
    monkeypatch.setattr(tools, "get_llm_client", lambda: RuleBasedLLMClient())
    client = TestClient(app)

    text = "\n".join(
        [
            "1. Ибупрофен 200мг №20 ЗЦ 45руб",
            "2. Парацетамол 500мг №10 ЗЦ 30руб",
            "3. Лоратадин 10мг №10 ЗЦ 80руб",
            "4. Амоксициллин 500мг №16 ЗЦ 120руб",
            "5. Цитрамон №10 ЗЦ 25руб",
        ]
    )
    parsed = client.post("/tools/parse_offer", json={"text": text}).json()
    assert len(parsed) == 5

    matched = []
    correct = 0
    for sku in parsed:
        esklp = client.post("/tools/search_esklp", json={"trade_name": sku["sku_name_raw"]}).json()
        enriched = dict(sku)
        if esklp:
            enriched.update(
                {
                    key: esklp[0].get(key)
                    for key in ("mnn", "form", "dosage", "atx_code", "atx_name", "ftg_name")
                }
            )
        match = client.post("/tools/match_pk", json={"trade_name": sku["sku_name_raw"], **enriched}).json()
        matched.append({"sku": enriched, "match": match})
        if match.get("status") == "auto_matched" and match.get("pk"):
            correct += 1

    assert correct >= 3

    with open("tests/fixtures/base_sample.xlsx", "rb") as file:
        base_file_id = client.post(
            "/tools/upload_base",
            files={"file": ("base_sample.xlsx", file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        ).json()["file_id"]

    build_response = client.post(
        "/tools/build_excel",
        json={"base_file_id": base_file_id, "matched_skus": matched},
    )
    assert build_response.status_code == 200

    download_response = client.get(build_response.json()["download_url"])
    assert download_response.status_code == 200
    output_path = tmp_path / "result.xlsx"
    output_path.write_bytes(download_response.content)
    workbook = load_workbook(output_path)
    assert "Сводка" in workbook.sheetnames
    assert "КП_распознано" in workbook.sheetnames